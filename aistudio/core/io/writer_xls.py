import  pandas as pd 
from    typing import List 
from    pathlib import Path 
from    openpyxl.styles import PatternFill, Font, Border, Side 


class ExcelFileWriter(object):
    """
    Example:
    filename_   = 'distribution-report.xlsx'
    sheet_name_ = 'admin'
    >>> xls    = ExcelFileWriter() 
    >>> df_xls = xls.read_excel(filename=filename_, sheet_name=sheet_name_).head()
    >>> xls.write_excel(dataset.data[['category', 'requested_by', 'volume']], filename_, sheet_name_, with_index=False)
    """
    def __init__(self, engine:str="openpyxl", **kwargs):
        self.engine = engine 
        self.kwargs = kwargs 
        # defaults: Create a pattern fill for coloring the header
        self.header_pattern = PatternFill(start_color="000033", end_color="000033", fill_type="solid") 
        self.pad_width      = 5 


    def read_excel(self, filename:str, sheet_name:str|List[str], **kwargs) -> pd.DataFrame:
        with pd.ExcelFile(filename) as reader:  
            self.filename   = filename 
            self.sheet_name = sheet_name
            return pd.read_excel(reader, sheet_name, engine=self.engine, **kwargs)  


    def read_sheet_names(self, filename:str, **kwargs) -> List[str]:
        with pd.ExcelFile(filename) as reader:
            return reader.sheet_names

    def write_excel(self, df:pd.DataFrame, filename:str, sheet_name:str, with_index:bool=True, **kwargs):
        # if file exists, mode='a' or 'w'
        mappings = dict(mode='a', if_sheet_exists='replace') if Path(filename).exists() else dict(mode='w')

        with pd.ExcelWriter(
            filename, 
            engine=self.engine, 
            date_format="YYYY-MM-DD",
            datetime_format="YYYY-MM-DD",
            **mappings
        ) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=with_index)
            # Access the XlsxWriter workbook and worksheet objects
            workbook   = writer.book
            sheets     = workbook.worksheets  
            worksheet  = writer.sheets[sheet_name]
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
            # adjust index properties first
            self.adjust_index_format(df, worksheet)
            # adjust header and widths
            self.adjust_header_cell(worksheet, header_row)
            self.adjust_widths(worksheet)
            # alignment wrapping 
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = cell.alignment.copy(wrapText=True)  # Enable wrap text for all cells

            workbook.save(filename)

    def adjust_header_cell(self, worksheet, header, freeze_panes:bool=True) -> None:
        # Color the header row (assuming first row is the header), but skipping the index values
        for cell in header:  
            cell.fill   = self.header_pattern
            cell.font   = Font(bold=True, color="FFFFFF")

        # Freeze the first row (A2 cell will be the top-left unfrozen cell), (above row 2)
        if freeze_panes:
            worksheet.freeze_panes = worksheet['A2']  

    def adjust_widths(self, worksheet, padding:int=5) -> None:
        # get column names 
        column_widths = {
            col[0].column_letter: max([len(str(cell.value)) if cell.value else 0 for cell in col]) + padding
            for col in worksheet.columns
        }
        # Set the column widths
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

    def adjust_index_format(self, frame:pd.DataFrame, worksheet) -> None: 
        # Define the style to remove bold and borders
        no_bold   = Font(bold=True, color="000000")
        no_border = Border()

        # Apply formatting to the index column (which is the first column) for the index values
        for i, idx in enumerate(frame.index, start=2):  # start=2 because row 1 is the header
            cell = worksheet.cell(row=i, column=1)      # Index column is always the first column
            cell.font   = no_bold
            cell.border = no_border

        # keep index name 
        if frame.index.name:
            index_name_cell        = worksheet.cell(row=1, column=1)
            index_name_cell.value  = frame.index.name  
            index_name_cell.font   = no_bold
            index_name_cell.border = no_border


        # # Write the index values without bold or border
        # for row_num, index_value in enumerate(frame.index, start=1):
        #     cell = worksheet.cell(row=row_num, column=index_value)
        #     cell.font   = n_bold 
        #     cell.border = no_border 
            

